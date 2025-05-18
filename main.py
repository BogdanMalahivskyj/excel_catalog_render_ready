from fastapi import FastAPI, File, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from PIL import Image
import io
import tempfile
from typing import List

app = FastAPI()

def cm_to_px(cm):
    return int(cm * 37.795275591)
    
@app.get("/")
def root():
    return {"status": "ok"}
    
@app.post("/generate_catalog")
async def generate_catalog(
    excel: UploadFile = File(...),
    images: List[UploadFile] = File(...)
):
    wb_data = load_workbook(excel.file)
    ws_data = wb_data.active
    rows = list(ws_data.iter_rows(min_row=2, values_only=True))

    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог"

    headers = ["#", "foto", "size", "description", "quantity", "price"]
    ws.append(headers)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 14.3
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15

    for i, (row_data, image_file) in enumerate(zip(rows, images), start=1):
        size, quantity, price = row_data
        description = f"Стіл з нержавіючої сталі розміром {size} мм. З нижньою полицею або мийкою (за фото)."
        row_idx = i + 1
        ws.row_dimensions[row_idx].height = 75

        ws[f"A{row_idx}"] = i
        ws[f"C{row_idx}"] = size
        ws[f"D{row_idx}"] = description
        ws[f"E{row_idx}"] = quantity
        ws[f"F{row_idx}"] = price

        for col in ["A", "C", "D", "E", "F"]:
            align = Alignment(horizontal="center", vertical="center", wrap_text=(col == "D"))
            ws[f"{col}{row_idx}"].alignment = align

        temp_img_path = tempfile.mktemp(suffix=".png")
        img = Image.open(image_file.file)
        img = img.resize((cm_to_px(2), cm_to_px(2)))
        img.save(temp_img_path)

        xl_img = XLImage(temp_img_path)
        xl_img.width = cm_to_px(2)
        xl_img.height = cm_to_px(2)
        xl_img.anchor = f"B{row_idx}"
        ws.add_image(xl_img)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=catalog_ready.xlsx"}
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
